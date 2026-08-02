#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт рецептов из Excel в menu.json для бота.

Запуск (в папке, где лежат папки кухонь Japan/Italy/Asia/Korea/SouthAmerica):
    python export_menu.py

Или с явными путями:
    python export_menu.py "C:\\Menu" "C:\\Menu\\menu-bot-addon\\menu_bot\\menu.json"

Требуется: pip install openpyxl
"""
import openpyxl, json, os, glob, sys, re

CUISINES = [
    ("Japan", "japan", "Япония"),
    ("Italy", "italy", "Италия"),
    ("Asia", "asia", "Тайланд / Азия"),
    ("Korea", "korea", "Корея"),
    ("SouthAmerica", "south_america", "Южная Америка"),
]

SOUP_KW = ["суп", "чиге", "том-кха", "минестроне", "мисо", "касуэла"]
DESSERT_KW = ["десерт", "моти", "тирамису", "хотток", "манго", "альфахорес"]


def find_xlsx(folder):
    for p in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        b = os.path.basename(p)
        if b.startswith("~$") or b.startswith(".~"):
            continue
        return p
    raise FileNotFoundError(f"Не найден .xlsx в папке: {folder}")


def dish_type(name):
    low = (name or "").lower()
    if any(k in low for k in SOUP_KW):
        return "суп"
    if any(k in low for k in DESSERT_KW):
        return "десерт"
    return "основное"


def age_num(age_str):
    m = re.search(r"\d+", str(age_str or ""))
    return int(m.group()) if m else 0


def dishes(folder):
    wb = openpyxl.load_workbook(find_xlsx(folder))
    out = []
    for ws in wb.worksheets:
        hdr = itg = cook = None
        for rr in range(1, ws.max_row + 1):
            a = ws.cell(rr, 1).value
            if ws.cell(rr, 4).value == "Ккал":
                hdr = rr
            if a == "ИТОГО, ккал":
                itg = rr
            if a == "ПРИГОТОВЛЕНИЕ":
                cook = rr
        if not hdr or not itg:
            continue
        ings = [
            {"name": str(ws.cell(rr, 2).value), "amount": str(ws.cell(rr, 3).value)}
            for rr in range(hdr + 1, itg)
        ]
        total = sum(ws.cell(rr, 4).value or 0 for rr in range(hdr + 1, itg))
        portions = ws["D2"].value or 1
        # шаги приготовления
        steps = []
        if cook:
            for rr in range(cook + 1, ws.max_row + 1):
                a = ws.cell(rr, 1).value
                b = ws.cell(rr, 2).value
                if a and str(a).startswith("Шаг") and b and str(b).strip():
                    steps.append(str(b).strip())
        photo = ws["D5"].value  # имя файла фото (или пусто)
        name = ws["A1"].value
        out.append({
            "name": name,
            "portions": portions,
            "time": ws["B3"].value,
            "spice": ws["B4"].value,
            "age": ws["D4"].value,
            "age_num": age_num(ws["D4"].value),
            "type": dish_type(name),
            "kcal_per_portion": round(total / portions),
            "photo": (str(photo).strip() if photo else ""),
            "ingredients": ings,
            "steps": steps,
        })
    wb.close()
    return out


def main():
    base = sys.argv[1] if len(sys.argv) > 1 else "."
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(base, "menu.json")
    menu = {"cuisines": []}
    for folder, key, title in CUISINES:
        path = os.path.join(base, folder)
        if not os.path.isdir(path):
            print(f"[!] Пропускаю (нет папки): {path}")
            continue
        menu["cuisines"].append({"key": key, "title": title, "dishes": dishes(path)})
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(menu, f, ensure_ascii=False, indent=2)
    n = sum(len(c["dishes"]) for c in menu["cuisines"])
    print(f"Готово: {out_path} — {len(menu['cuisines'])} кухонь, {n} блюд")


if __name__ == "__main__":
    main()
